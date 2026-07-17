#!/usr/bin/env python3
"""
Motor de Consolidacion de IVA (CALA) — VERSION 2, dos pasos.

MODO PREVIEW:
    python3 consolidar_iva.py preview <listado.xlsx> [--tarifas <json_nit_tarifa>]
    -> devuelve JSON con todas las facturas, su tarifa presunta y el cliente detectado.

MODO GENERAR:
    python3 consolidar_iva.py generar <listado.xlsx> <salida.xlsx> <decisiones_json>
    -> decisiones_json: ruta a un JSON { cufe: tarifa }. Genera el Excel.

Requiere: pip install openpyxl --break-system-packages
"""
import sys
import io
import json
import unicodedata
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# ---- Columnas del listado DIAN (Rp_Doc), 0-based ----
C_TIPO, C_CUFE, C_FOLIO, C_PREFIJO = 0, 1, 2, 3
C_FORMA_PAGO = 5
C_FECHA_EMI = 7
C_NIT_EMI, C_NOM_EMI, C_NIT_REC, C_NOM_REC = 9, 10, 11, 12
C_IVA = 13
C_RETE_IVA, C_RETE_RENTA, C_RETE_ICA = 26, 27, 28
C_TOTAL, C_ESTADO, C_GRUPO = 29, 30, 31
C_OTROS = list(range(14, 26))

ENCABEZADOS = [
    "Tipo de documento", "CUFE/CUDE", "Folio", "Prefijo", "Fecha Emisión",
    "NIT Emisor", "Nombre Emisor", "NIT Receptor", "Nombre Receptor", "Forma Pago",
    "Base Exenta", "Base Gravada 19%", "Base Gravada 5%",
    "IVA 19%", "IVA 5%", "IVA Total", "Otros Impuestos",
    "Rete IVA", "Rete Renta", "Rete ICA",
    "TOTAL Calculado", "Total DIAN", "Diferencia", "Tarifa", "Estado",
]
COL = {name: get_column_letter(i + 1) for i, name in enumerate(ENCABEZADOS)}
NEGRITA = Font(bold=True, color="FFFFFF")
RELLENO_HEAD = PatternFill("solid", fgColor="1F4E78")
RELLENO_TOT = PatternFill("solid", fgColor="DDEBF7")
MONEDA = '#,##0;(#,##0);"-"'


def _norm(s):
    s = (s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def clasificar(tipo, grupo):
    t, g = _norm(tipo), _norm(grupo)
    if "application response" in t:
        return None, False
    es_nc = "nota" in t and "credito" in t
    if g == "emitido":
        if "nomina" in t:
            return "NOMINA", False
        if "documento soporte" in t:
            return "DOCUMENTO SOPORTE", False
        return "VENTAS", es_nc
    if g == "recibido":
        return "COMPRAS", es_nc
    return None, False


def fnum(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _celda(fila, idx):
    return fila[idx] if idx < len(fila) else None


def leer_facturas(listado_path):
    wb = load_workbook(listado_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))

    inicio = 1
    for i, fila in enumerate(filas[:15]):
        textos = [_norm(str(c)) for c in fila if c is not None]
        if any("cufe" in t or "cude" in t for t in textos):
            inicio = i + 1
            break

    out = []
    for fila in filas[inicio:]:
        if fila is None or _celda(fila, C_TIPO) is None:
            continue
        hoja, es_nc = clasificar(_celda(fila, C_TIPO), _celda(fila, C_GRUPO))
        if hoja is None:
            continue
        out.append({
            "cufe": str(_celda(fila, C_CUFE) or ""),
            "hoja": hoja, "es_nc": es_nc,
            "tipo": str(_celda(fila, C_TIPO) or ""),
            "folio": _celda(fila, C_FOLIO),
            "prefijo": _celda(fila, C_PREFIJO),
            "fecha": str(_celda(fila, C_FECHA_EMI) or ""),
            "nit_emi": str(_celda(fila, C_NIT_EMI) or ""),
            "nom_emi": str(_celda(fila, C_NOM_EMI) or ""),
            "nit_rec": str(_celda(fila, C_NIT_REC) or ""),
            "nom_rec": str(_celda(fila, C_NOM_REC) or ""),
            "forma": str(_celda(fila, C_FORMA_PAGO) or ""),
            "iva_dian": fnum(_celda(fila, C_IVA)),
            "otros": sum(fnum(_celda(fila, i)) for i in C_OTROS),
            "rete_iva": fnum(_celda(fila, C_RETE_IVA)),
            "rete_renta": fnum(_celda(fila, C_RETE_RENTA)),
            "rete_ica": fnum(_celda(fila, C_RETE_ICA)),
            "total_dian": fnum(_celda(fila, C_TOTAL)),
            "estado": str(_celda(fila, C_ESTADO) or ""),
        })
    return out


def _nit_prov(f):
    return f["nit_emi"] if f["hoja"] == "COMPRAS" else f["nit_rec"]


def _nom_prov(f):
    return f["nom_emi"] if f["hoja"] == "COMPRAS" else f["nom_rec"]


def _detectar_cliente(facturas):
    """
    El cliente (la empresa consolidada) es el emisor en las ventas/nomina/doc
    soporte, y el receptor en las compras. Se toma el NIT mas frecuente.
    """
    cont = Counter()
    nombres = {}
    for f in facturas:
        if f["hoja"] == "COMPRAS":
            nit, nom = f["nit_rec"], f["nom_rec"]
        else:
            nit, nom = f["nit_emi"], f["nom_emi"]
        if nit:
            cont[nit] += 1
            nombres.setdefault(nit, nom)
    if not cont:
        return "", ""
    nit = cont.most_common(1)[0][0]
    return nit, nombres.get(nit, "")


def modo_preview(listado_path, tarifas_conocidas):
    facturas = leer_facturas(listado_path)
    cliente_nit, cliente_nombre = _detectar_cliente(facturas)
    filas = []
    for f in facturas:
        nit = _nit_prov(f)
        if nit in tarifas_conocidas:
            tarifa, origen = tarifas_conocidas[nit], "conocido"
        else:
            tarifa, origen = 19, "presunto"
        filas.append({
            "cufe": f["cufe"], "hoja": f["hoja"], "tipo": f["tipo"],
            "fecha": f["fecha"], "nit_proveedor": nit,
            "nombre_proveedor": _nom_prov(f),
            "iva_dian": f["iva_dian"], "total_dian": f["total_dian"],
            "es_nc": f["es_nc"], "tarifa": tarifa, "origen": origen,
        })
    return {
        "total": len(filas),
        "facturas": filas,
        "cliente_nit": cliente_nit,
        "cliente_nombre": cliente_nombre,
    }


def _bases(f, tarifa):
    iva, total, otros = f["iva_dian"], f["total_dian"], f["otros"]
    base_ex = base19 = base5 = iva19 = iva5 = 0.0
    if tarifa == 19:
        iva19 = iva
        base19 = round(iva / 0.19, 2) if iva else 0.0
        base_ex = round(total - base19 - iva19 - otros, 2)
    elif tarifa == 5:
        iva5 = iva
        base5 = round(iva / 0.05, 2) if iva else 0.0
        base_ex = round(total - base5 - iva5 - otros, 2)
    elif tarifa == 0:
        base_ex = round(total - otros, 2)
    else:
        t = tarifa / 100.0
        if t > 0 and iva:
            base19 = round(iva / t, 2)
            iva19 = iva
            base_ex = round(total - base19 - iva - otros, 2)
        else:
            base_ex = round(total - otros, 2)
    return base_ex, base19, base5, iva19, iva5


def modo_generar(listado_path, salida_path, decisiones):
    facturas = leer_facturas(listado_path)
    hojas = {"VENTAS": [], "COMPRAS": [], "NOMINA": [], "DOCUMENTO SOPORTE": []}
    resumen = {"total_filas": 0, "por_tarifa": {}}
    for f in facturas:
        resumen["total_filas"] += 1
        tarifa = decisiones.get(f["cufe"], 19)
        k = "Exento" if tarifa == 0 else f"{tarifa}%"
        resumen["por_tarifa"][k] = resumen["por_tarifa"].get(k, 0) + 1
        be, b19, b5, i19, i5 = _bases(f, tarifa)
        s = -1 if f["es_nc"] else 1
        hojas[f["hoja"]].append({
            **f, "base_ex": s*be, "base19": s*b19, "base5": s*b5,
            "iva19": s*i19, "iva5": s*i5, "otros_s": s*f["otros"],
            "rete_iva_s": s*f["rete_iva"], "rete_renta_s": s*f["rete_renta"],
            "rete_ica_s": s*f["rete_ica"], "total_s": s*f["total_dian"], "tarifa": tarifa,
        })
    wb = Workbook()
    wb.remove(wb.active)
    totales = {}
    for nombre in ["VENTAS", "COMPRAS", "NOMINA", "DOCUMENTO SOPORTE"]:
        totales[nombre] = _escribir_hoja(wb, nombre, hojas[nombre])
    _escribir_acumulado(wb, totales)
    wb.save(salida_path)
    return resumen


def _escribir_hoja(wb, nombre, registros):
    ws = wb.create_sheet(nombre)
    for j, h in enumerate(ENCABEZADOS, start=1):
        c = ws.cell(1, j, h)
        c.font = NEGRITA; c.fill = RELLENO_HEAD
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    fila = 2
    for r in registros:
        vals = [r["tipo"], r["cufe"], r["folio"], r["prefijo"], r["fecha"],
                r["nit_emi"], r["nom_emi"], r["nit_rec"], r["nom_rec"], r["forma"],
                r["base_ex"], r["base19"], r["base5"], r["iva19"], r["iva5"]]
        for j, v in enumerate(vals, start=1):
            ws.cell(fila, j, v)
        ws.cell(fila, 16, f"=N{fila}+O{fila}")
        ws.cell(fila, 17, r["otros_s"])
        ws.cell(fila, 18, r["rete_iva_s"])
        ws.cell(fila, 19, r["rete_renta_s"])
        ws.cell(fila, 20, r["rete_ica_s"])
        ws.cell(fila, 21, f"=K{fila}+L{fila}+M{fila}+P{fila}+Q{fila}")
        ws.cell(fila, 22, r["total_s"])
        ws.cell(fila, 23, f"=V{fila}-U{fila}")
        ws.cell(fila, 24, "Exento" if r["tarifa"] == 0 else f'{r["tarifa"]}%')
        ws.cell(fila, 25, r["estado"])
        for col in [11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23]:
            ws.cell(fila, col).number_format = MONEDA
        fila += 1
    if fila > 2:
        ws.cell(fila, 10, "TOTALES").font = Font(bold=True)
        for cidx in [11, 12, 13, 14, 15, 16, 21, 22]:
            letra = get_column_letter(cidx)
            c = ws.cell(fila, cidx, f"=SUM({letra}2:{letra}{fila-1})")
            c.font = Font(bold=True); c.fill = RELLENO_TOT; c.number_format = MONEDA
    anchos = [18, 30, 8, 8, 12, 14, 24, 14, 24, 12, 14, 14, 14, 14, 14,
              14, 12, 12, 12, 12, 16, 16, 12, 10, 12]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return fila


def _escribir_acumulado(wb, totales):
    ws = wb.create_sheet("ACUMULADO", 0)
    ws.cell(2, 2, "CONSOLIDACIÓN DE IVA").font = Font(bold=True, size=14)
    def ref(hoja, col):
        return f"='{hoja}'!{COL[col]}{totales[hoja]}"
    headers = ["Concepto", "Base Exenta", "Base 19%", "Base 5%", "IVA 19%", "IVA 5%", "IVA Total"]
    fila = 4
    for j, h in enumerate(headers, start=2):
        c = ws.cell(fila, j, h); c.font = NEGRITA; c.fill = RELLENO_HEAD
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    fila += 1
    for etiqueta, hoja in [("VENTAS", "VENTAS"), ("COMPRAS", "COMPRAS")]:
        ws.cell(fila, 2, etiqueta).font = Font(bold=True)
        ws.cell(fila, 3, ref(hoja, "Base Exenta"))
        ws.cell(fila, 4, ref(hoja, "Base Gravada 19%"))
        ws.cell(fila, 5, ref(hoja, "Base Gravada 5%"))
        ws.cell(fila, 6, ref(hoja, "IVA 19%"))
        ws.cell(fila, 7, ref(hoja, "IVA 5%"))
        ws.cell(fila, 8, ref(hoja, "IVA Total"))
        for j in range(3, 9):
            ws.cell(fila, j).number_format = MONEDA
        fila += 1
    fila += 1
    ws.cell(fila, 2, "IVA A PAGAR (Ventas - Compras)").font = Font(bold=True, size=12)
    ws.cell(fila, 8, "=H5-H6"); ws.cell(fila, 8).number_format = MONEDA
    ws.cell(fila, 8).font = Font(bold=True, size=12)
    ws.cell(fila, 8).fill = PatternFill("solid", fgColor="FFF2CC")
    for i, w in enumerate([3, 34, 16, 16, 16, 14, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


if __name__ == "__main__":
    modo = sys.argv[1] if len(sys.argv) > 1 else ""
    if modo == "preview":
        listado = sys.argv[2]
        tarifas = {}
        if "--tarifas" in sys.argv:
            tarifas = json.loads(sys.argv[sys.argv.index("--tarifas") + 1])
        print(json.dumps(modo_preview(listado, tarifas), ensure_ascii=False, default=str))
    elif modo == "generar":
        listado, salida, dec_path = sys.argv[2], sys.argv[3], sys.argv[4]
        with open(dec_path, encoding="utf-8") as fh:
            decisiones = json.load(fh)
        decisiones = {k: int(v) for k, v in decisiones.items()}
        res = modo_generar(listado, salida, decisiones)
        print(json.dumps(res, ensure_ascii=False, default=str))
    else:
        print(json.dumps({"error": f"Modo desconocido: {modo}"}), file=sys.stderr)
        sys.exit(1)