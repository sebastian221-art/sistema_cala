# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

"""
generador_1647.py
Procesa un archivo AUXILIAR (.xls de Siigo/World Office) y genera el Formulario 1647
(Ingresos recibidos para terceros) en formato Excel.
"""

import sys
import xlrd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from collections import defaultdict


# ─── Helpers ──────────────────────────────────────────────────────────────────

def calcular_dv(nit_raw) -> str:
    """Calcula el dígito de verificación de un NIT colombiano."""
    try:
        nit = str(int(float(nit_raw))).strip()
    except Exception:
        nit = str(nit_raw).strip().split('.')[0]
    if not nit or not nit.isdigit():
        return ''
    weights = [71, 67, 59, 53, 47, 43, 41, 37, 29, 23, 19, 17, 13, 7, 3]
    nit_padded = nit.zfill(15)
    total = sum(int(d) * w for d, w in zip(nit_padded, weights))
    r = total % 11
    if r == 0:
        return '0'
    elif r == 1:
        return '1'
    else:
        return str(11 - r)


def limpiar_nit(nit_raw) -> str:
    """Convierte el NIT a string limpio sin decimales."""
    try:
        return str(int(float(nit_raw)))
    except Exception:
        return str(nit_raw).strip().split('.')[0]


# DESPUÉS
def split_nombre(nombre: str, tipdoc: str, nit: str = '') -> dict:
    nombre = str(nombre).strip()
    # tipdoc 31 = empresa local, 44444xxx = empresa extranjera → razón social
    if tipdoc == '31' or nit.startswith('44444'):
        return {
            'primer_apellido': '',
            'segundo_apellido': '',
            'primer_nombre': '',
            'otros_nombres': '',
            'razon_social': nombre,
        }
    partes = [p for p in nombre.split() if p not in ('Y/O', '-')]
    return {
        'primer_apellido': partes[0] if len(partes) > 0 else '',
        'segundo_apellido': partes[1] if len(partes) > 1 else '',
        'primer_nombre': partes[2] if len(partes) > 2 else '',
        'otros_nombres': ' '.join(partes[3:]) if len(partes) > 3 else '',
        'razon_social': '',
    }

def codigo_pais(tipdoc: str) -> str:
    """Código de país: 169 = Colombia para tipos de doc locales."""
    extranjeros = {'43', '22', '41', '42'}
    return '' if tipdoc in extranjeros else '169'


# ─── Leer AUXILIAR ────────────────────────────────────────────────────────────

def leer_auxiliar(path: str) -> list[dict]:
    """Lee el archivo AUXILIAR .xls y retorna lista de filas como dicts."""
    wb = xlrd.open_workbook(path)
    sh = wb.sheets()[0]
    headers = [sh.cell_value(0, c) for c in range(sh.ncols)]

    filas = []
    for r in range(1, sh.nrows):
        fila = {headers[c]: sh.cell_value(r, c) for c in range(sh.ncols)}
        filas.append(fila)
    return filas


# ─── Procesar datos ───────────────────────────────────────────────────────────

def procesar_auxiliar(filas: list[dict]) -> list[dict]:
    """
    Transforma las filas del AUXILIAR en registros del Formulario 1647.

    Lógica:
    - Agrupa por 'numero' (documento G).
    - Filas con debito > 0 -> lado izquierdo del 1647 (cols C-L, K).
    - Filas con credito > 0 -> lado derecho del 1647 (cols O-X).
    - K = total debito acumulado por tercero pagador (col M).
    - L = valor individual de cada credito.
    - Una fila del 1647 por cada par (debito_row, credito_row) del mismo documento.
    """

    # Solo acumula debitos de docs que TAMBIEN tienen credito
    docs_temp: dict = defaultdict(lambda: {'debito': [], 'credito': []})
    for f in filas:
        numero = str(f.get('numero', '')).strip()
        deb = f.get('debito', 0) or 0
        cre = f.get('credito', 0) or 0
        if float(deb) > 0:
            docs_temp[numero]['debito'].append(f)
        if float(cre) > 0:
            docs_temp[numero]['credito'].append(f)

    total_debito_por_tercero: dict = defaultdict(float)
    for numero, data in docs_temp.items():
        if data['debito'] and data['credito']:
            for f in data['debito']:
                tercero = str(f.get('tercero', '')).strip()
                total_debito_por_tercero[tercero] += float(f.get('debito', 0) or 0)

    # Agrupar por numero de documento
    docs: dict[str, dict] = defaultdict(lambda: {'debito': [], 'credito': []})
    for f in filas:
        numero = str(f.get('numero', '')).strip()
        deb = f.get('debito', 0) or 0
        cre = f.get('credito', 0) or 0
        if float(deb) > 0:
            docs[numero]['debito'].append(f)
        if float(cre) > 0:
            docs[numero]['credito'].append(f)

    registros: list[dict] = []

    for numero, data in docs.items():
        debs = data['debito']
        cres = data['credito']

        if not debs or not cres:
            continue

        pag = debs[0]
        nit_pag    = limpiar_nit(pag.get('nit', ''))
        tip_pag    = str(pag.get('tipdoc', '')).strip()
        ter_pag    = str(pag.get('tercero', '')).strip()
        dv_pag     = calcular_dv(nit_pag)
        pais_pag   = codigo_pais(tip_pag)
        k_valor    = total_debito_por_tercero.get(ter_pag, 0)
        nombre_pag = split_nombre(ter_pag, tip_pag, nit_pag)

        for rec in cres:
            nit_rec    = limpiar_nit(rec.get('nit', ''))
            tip_rec    = str(rec.get('tipdoc', '')).strip()
            ter_rec    = str(rec.get('tercero', '')).strip()
            dv_rec     = calcular_dv(nit_rec)
            pais_rec   = codigo_pais(tip_rec)
            l_valor    = float(rec.get('credito', 0) or 0)
            nombre_rec = split_nombre(ter_rec, tip_rec, nit_rec)
            depto      = str(rec.get('depto_ter', '') or rec.get('depto', '') or '').strip()
            muni       = str(rec.get('muni_ter',  '') or rec.get('muni',  '') or '').strip()

            registros.append({
                'concepto':          'A070',
                'tip_doc_pag':       tip_pag,
                'nit_pag':           nit_pag,
                'dv_pag':            dv_pag,
                'pap1_pag':          nombre_pag['primer_apellido'],
                'pap2_pag':          nombre_pag['segundo_apellido'],
                'pnom_pag':          nombre_pag['primer_nombre'],
                'otros_pag':         nombre_pag['otros_nombres'],
                'razon_pag':         nombre_pag['razon_social'],
                'pais_pag':          pais_pag,
                'valor_total':       k_valor,
                'valor_transferido': l_valor,
                'tip_doc_rec':       tip_rec,
                'nit_rec':           nit_rec,
                'dv_rec':            dv_rec,
                'pap1_rec':          nombre_rec['primer_apellido'],
                'pap2_rec':          nombre_rec['segundo_apellido'],
                'pnom_rec':          nombre_rec['primer_nombre'],
                'otros_rec':         nombre_rec['otros_nombres'],
                'razon_rec':         nombre_rec['razon_social'],
                'dir_rec':           '',
                'depto_rec':         depto,
                'muni_rec':          muni,
                'pais_rec':          pais_rec,
            })

    return registros


# ─── Generar Excel ────────────────────────────────────────────────────────────

# Colores
AZUL_HEADER  = 'FF003366'  # azul oscuro DIAN
AZUL_SUB     = 'FF336699'  # azul medio
GRIS_HEADER  = 'FFD9D9D9'
BLANCO       = 'FFFFFFFF'
AMARILLO     = 'FFFFFF00'

def thin_border():
    thin = Side(style='thin', color='FF000000')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def estilo_header(ws, cell_ref: str, texto: str,
                  bold=True, color_bg=AZUL_HEADER, color_font=BLANCO,
                  wrap=True, h_align='center'):
    cell = ws[cell_ref]
    cell.value = texto
    cell.font = Font(bold=bold, color=color_font, size=8,
                     name='Arial')
    cell.fill = PatternFill('solid', start_color=color_bg)
    cell.alignment = Alignment(horizontal=h_align, vertical='center',
                                wrap_text=wrap)
    cell.border = thin_border()


def generar_excel(registros: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1647"

    # ── Fila 1: Título ─────────────────────────────────────────────────────
    ws.merge_cells('A1:X1')
    c = ws['A1']
    c.value = 'FORMULARIO 1647 — INGRESOS RECIBIDOS PARA TERCEROS'
    c.font = Font(bold=True, size=11, color=BLANCO, name='Arial')
    c.fill = PatternFill('solid', start_color=AZUL_HEADER)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # ── Fila 2: sub-encabezados de bloque ──────────────────────────────────
    ws.merge_cells('A2:A4')
    estilo_header(ws, 'A2', 'Concepto', color_bg=AZUL_SUB)

    ws.merge_cells('B2:J2')
    estilo_header(ws, 'B2', 'DATOS DE QUIEN RECIBE EL INGRESO', color_bg=AZUL_HEADER)

    ws.merge_cells('K2:N2')
    estilo_header(ws, 'K2', 'VALORES', color_bg=AZUL_HEADER)

    ws.merge_cells('O2:X2')
    estilo_header(ws, 'O2', 'DATOS DEL TERCERO PARA QUIEN SE RECIBIÓ EL INGRESO', color_bg=AZUL_HEADER)

    # ── Fila 3: sub-bloques ────────────────────────────────────────────────
    ws.merge_cells('B3:B4')
    estilo_header(ws, 'B3', 'Tipo\ndoc', color_bg=AZUL_SUB, color_font=BLANCO)

    ws.merge_cells('C3:C4')
    estilo_header(ws, 'C3', 'NIT / Identificación', color_bg=AZUL_SUB)

    ws.merge_cells('D3:D4')
    estilo_header(ws, 'D3', 'DV', color_bg=AZUL_SUB)

    ws.merge_cells('E3:I3')
    estilo_header(ws, 'E3', 'Nombre de quien se recibe el ingreso', color_bg=AZUL_SUB)

    ws.merge_cells('J3:J4')
    estilo_header(ws, 'J3', 'País\nresidencia', color_bg=AZUL_SUB)

    ws.merge_cells('K3:K4')
    estilo_header(ws, 'K3', 'Valor total\nde la\noperación', color_bg=AZUL_SUB)

    ws.merge_cells('L3:L4')
    estilo_header(ws, 'L3', 'Valor ingreso\nreintegrado /\ntransferido', color_bg=AZUL_SUB)

    ws.merge_cells('M3:M4')
    estilo_header(ws, 'M3', 'Tipo\nretención', color_bg=AZUL_SUB)

    ws.merge_cells('N3:N4')
    estilo_header(ws, 'N3', 'Valor\nretención\ntransferida', color_bg=AZUL_SUB)

    ws.merge_cells('O3:O4')
    estilo_header(ws, 'O3', 'Tipo\ndoc', color_bg=AZUL_SUB)

    ws.merge_cells('P3:P4')
    estilo_header(ws, 'P3', 'NIT / Identificación', color_bg=AZUL_SUB)

    ws.merge_cells('Q3:Q4')
    estilo_header(ws, 'Q3', 'DV', color_bg=AZUL_SUB)

    ws.merge_cells('R3:V3')
    estilo_header(ws, 'R3', 'Nombre del tercero para quien se recibió el ingreso', color_bg=AZUL_SUB)

    ws.merge_cells('W3:W4')
    estilo_header(ws, 'W3', 'Cód\ndpto', color_bg=AZUL_SUB)

    ws.merge_cells('X3:X4')
    estilo_header(ws, 'X3', 'Cód\nmpio', color_bg=AZUL_SUB)

    # ── Fila 4: etiquetas de columnas de nombre ───────────────────────────
    nombre_cols_pag = {
        'E4': 'Primer apellido',
        'F4': 'Segundo apellido',
        'G4': 'Primer nombre',
        'H4': 'Otros nombres',
        'I4': 'Razón social',
    }
    for ref, txt in nombre_cols_pag.items():
        estilo_header(ws, ref, txt, color_bg=GRIS_HEADER, color_font='FF000000')

    nombre_cols_rec = {
        'R4': 'Primer apellido',
        'S4': 'Segundo apellido',
        'T4': 'Primer nombre',
        'U4': 'Otros nombres',
        'V4': 'Razón social',
    }
    for ref, txt in nombre_cols_rec.items():
        estilo_header(ws, ref, txt, color_bg=GRIS_HEADER, color_font='FF000000')

    # Completar celdas vacías de fila 4 con borde
    for col in ['A', 'B', 'C', 'D', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'W', 'X']:
        cell = ws[f'{col}4']
        if not cell.value:
            cell.fill = PatternFill('solid', start_color=GRIS_HEADER)
            cell.border = thin_border()

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 35
    ws.row_dimensions[4].height = 30

    # ── Datos ──────────────────────────────────────────────────────────────
    DATA_FILL   = PatternFill('solid', start_color='FFF2F2F2')
    DATA_FILL_B = PatternFill('solid', start_color=BLANCO)
    NUM_FONT    = Font(name='Arial', size=8)
    TXT_FONT    = Font(name='Arial', size=8)
    NUM_FMT     = '#,##0'

    for idx, reg in enumerate(registros):
        row = idx + 5
        fill = DATA_FILL if idx % 2 == 0 else DATA_FILL_B

        def w(col, val, num_format=None, align='left'):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal=align, vertical='center')
            cell.font = NUM_FONT if num_format else TXT_FONT
            if num_format:
                cell.number_format = num_format

        w(1,  reg['concepto'])                        # A: Concepto
        w(2,  reg['tip_doc_pag'], align='center')     # B: Tipo doc pagador
        w(3,  reg['nit_pag'])                         # C: NIT pagador
        w(4,  reg['dv_pag'], align='center')          # D: DV
        w(5,  reg['pap1_pag'])                        # E: Primer apellido
        w(6,  reg['pap2_pag'])                        # F: Segundo apellido
        w(7,  reg['pnom_pag'])                        # G: Primer nombre
        w(8,  reg['otros_pag'])                       # H: Otros nombres
        w(9,  reg['razon_pag'])                       # I: Razón social
        w(10, reg['pais_pag'], align='center')        # J: País
        w(11, reg['valor_total'] if reg['valor_total'] else None,
              num_format=NUM_FMT, align='right')      # K: Valor total
        w(12, reg['valor_transferido'] if reg['valor_transferido'] else None,
              num_format=NUM_FMT, align='right')      # L: Valor transferido
        w(13, '', align='center')                     # M: Tipo retención (vacío)
        w(14, None, num_format=NUM_FMT, align='right')# N: Val retención (vacío)
        w(15, reg['tip_doc_rec'], align='center')     # O: Tipo doc receptor
        w(16, reg['nit_rec'])                         # P: NIT receptor
        w(17, reg['dv_rec'], align='center')          # Q: DV receptor
        w(18, reg['pap1_rec'])                        # R: Primer apellido
        w(19, reg['pap2_rec'])                        # S: Segundo apellido
        w(20, reg['pnom_rec'])                        # T: Primer nombre
        w(21, reg['otros_rec'])                       # U: Otros nombres
        w(22, reg['razon_rec'])                       # V: Razón social
        w(23, reg['depto_rec'], align='center')       # W: Cód dpto
        w(24, reg['muni_rec'], align='center')        # X: Cód mpio

    # ── Anchos de columna ──────────────────────────────────────────────────
    anchos = {
        'A': 8, 'B': 6, 'C': 14, 'D': 4, 'E': 16, 'F': 16,
        'G': 14, 'H': 12, 'I': 28, 'J': 6, 'K': 14, 'L': 14,
        'M': 6, 'N': 14, 'O': 6, 'P': 14, 'Q': 4, 'R': 16,
        'S': 16, 'T': 14, 'U': 12, 'V': 28, 'W': 6, 'X': 6,
    }
    for col_letter, ancho in anchos.items():
        ws.column_dimensions[col_letter].width = ancho

    # ── Fila de totales ────────────────────────────────────────────────────
    total_row = len(registros) + 5
    ws.cell(total_row, 1).value = 'TOTALES'
    ws.cell(total_row, 1).font  = Font(bold=True, name='Arial', size=8)
    ws.cell(total_row, 1).fill  = PatternFill('solid', start_color=GRIS_HEADER)
    ws.cell(total_row, 1).border = thin_border()

    for col in range(2, 25):
        cell = ws.cell(total_row, col)
        cell.fill   = PatternFill('solid', start_color=GRIS_HEADER)
        cell.border = thin_border()
        cell.font   = Font(bold=True, name='Arial', size=8)

    # Suma de K y L
    if len(registros) > 0:
        r_ini = 5
        r_fin = total_row - 1
        ws.cell(total_row, 11).value  = f'=SUM(K{r_ini}:K{r_fin})'
        ws.cell(total_row, 11).number_format = NUM_FMT
        ws.cell(total_row, 11).alignment = Alignment(horizontal='right')

        ws.cell(total_row, 12).value  = f'=SUM(L{r_ini}:L{r_fin})'
        ws.cell(total_row, 12).number_format = NUM_FMT
        ws.cell(total_row, 12).alignment = Alignment(horizontal='right')

    # Congelar encabezados
    ws.freeze_panes = 'A5'

    wb.save(output_path)
    print(f'[OK] Generado: {output_path} — {len(registros)} registros')


# ─── Punto de entrada ─────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print('Uso: python generador_1647.py <auxiliar.xls> <salida_1647.xlsx>')
        sys.exit(1)

    auxiliar_path = sys.argv[1]
    output_path   = sys.argv[2]

    print(f'Leyendo AUXILIAR: {auxiliar_path}')
    filas = leer_auxiliar(auxiliar_path)
    print(f'  -> {len(filas)} filas leidas')

    print('Procesando...')
    registros = procesar_auxiliar(filas)
    print(f'  -> {len(registros)} registros para el 1647')

    print('Generando Excel...')
    generar_excel(registros, output_path)


if __name__ == '__main__':
    main()